#!/usr/bin/env /usr/bin/python3
"""Build the BGEP financial model as a LIVE Excel workbook for Raemy.

She asked for "the model in excel format" on 11 Aug 2026. A CFO asking for Excel is not asking
for a picture of the numbers, she is asking to change an assumption and watch the answer move.
So every output cell here is a FORMULA off the Assumptions sheet, not a pasted value.

Four sheets:
  Assumptions   - the only place anything is typed. Change a yellow cell, everything follows.
  Unit          - the walk from shelf price to our contribution per sachet
  Scenarios     - the downside grid, margin cases crossed with volume
  Setup         - one-off cost to get going

Run with SYSTEM python: openpyxl is on /usr/bin/python3 only. See ops_shared_tools_need_system_python.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ORANGE = "E0521F"
INPUT_FILL = PatternFill("solid", fgColor="FFF2CC")     # yellow = you may type here
HEAD_FILL = PatternFill("solid", fgColor="F0ECE3")
BOLD = Font(bold=True)
WHITE_BOLD = Font(bold=True, color="FFFFFF")
TITLE = Font(bold=True, size=14, color=ORANGE)
thin = Side(style="thin", color="D9D9D9")
BOX = Border(left=thin, right=thin, top=thin, bottom=thin)

wb = Workbook()


def title(ws, text, sub=""):
    ws["A1"] = text
    ws["A1"].font = TITLE
    if sub:
        ws["A2"] = sub
        ws["A2"].font = Font(italic=True, color="6F685C")
    ws.freeze_panes = "A4"


# ---------------------------------------------------------------- ASSUMPTIONS
a = wb.active
a.title = "Assumptions"
title(a, "BEAR GRYLLS EVEREST POWER — ASSUMPTIONS",
      "Type only in the yellow cells. Every other sheet is formulas off this one.")

rows = [
    ("FX", "", "", ""),
    ("INR per GBP", 128.70, "ECB reference rate, 10 Aug 2026, verified at source", "fx"),
    ("INR per USD", 95.30, "ECB reference rate, 10 Aug 2026", ""),
    ("", "", "", ""),
    ("PRICE AND CHANNEL", "", "", ""),
    ("Consumer price, Rs (MRP)", 10.00, "The India impulse price point", "mrp"),
    ("GST rate", 0.18, "Proprietary food, HSN 2106, no caffeine", "gst"),
    ("Retailer margin", 0.15, "Kirana FMCG band is 8-15%; top of band to buy shelf", "ret"),
    ("Distributor margin", 0.08, "District distributor norm is 4-8%. ASSUMED, not agreed", "dis"),
    ("", "", "", ""),
    ("COST PER SACHET, Rs", "", "", ""),
    ("Powder", 1.15, "Dextrose, salts, acidulant, flavour, vitamin C", "c1"),
    ("Sachet laminate and fill", 0.55, "", "c2"),
    ("Outer box of 100 and carton", 0.22, "", "c3"),
    ("Inward freight and wastage", 0.18, "", "c4"),
    ("Secondary freight to distributor", 0.15, "", "frt"),
    ("Trade spend per sachet", 0.30, "Distributor incentives, scheme, app rewards", "trd"),
    ("", "", "", ""),
    ("VOLUME", "", "", ""),
    ("Outlets, year 1", 60000, "0.5% of India's 13m kirana shops", "out"),
    ("Sachets per outlet per day", 2.5, "CORRECTED 10 Aug. The pitch assumed 10, which is ~4x too high", "rate"),
    ("Days per month", 30, "", "days"),
]
a["A4"], a["B4"], a["C4"] = "Assumption", "Value", "Note"
for c in ("A4", "B4", "C4"):
    a[c].font = BOLD
    a[c].fill = HEAD_FILL
    a[c].border = BOX

r = 5
names = {}
for label, val, note, key in rows:
    if label and val == "":
        a.cell(r, 1, label).font = Font(bold=True, color=ORANGE)
    elif label:
        a.cell(r, 1, label)
        cell = a.cell(r, 2, val)
        cell.fill = INPUT_FILL
        cell.border = BOX
        if isinstance(val, float) and val < 1:
            cell.number_format = "0.00%" if key in ("gst", "ret", "dis") else "0.00"
        elif isinstance(val, float):
            cell.number_format = "0.00"
        else:
            cell.number_format = "#,##0"
        a.cell(r, 3, note).font = Font(color="6F685C", size=9)
        if key:
            names[key] = f"Assumptions!$B${r}"
    r += 1

a.column_dimensions["A"].width = 34
a.column_dimensions["B"].width = 14
a.column_dimensions["C"].width = 62

# ---------------------------------------------------------------- UNIT
u = wb.create_sheet("Unit")
title(u, "UNIT ECONOMICS, Rs PER SACHET",
      "Every figure is a formula. Change an assumption and this moves.")
u["A4"], u["B4"] = "Line", "Rs"
for c in ("A4", "B4"):
    u[c].font = BOLD
    u[c].fill = HEAD_FILL
    u[c].border = BOX

unit = [
    ("Consumer pays (MRP)", f"={names['mrp']}"),
    ("Net of GST", f"={names['mrp']}/(1+{names['gst']})"),
    ("Retailer buys at", f"=B6*(1-{names['ret']})"),
    ("Distributor buys at — WE RECEIVE THIS", f"=B7*(1-{names['dis']})"),
    ("Less cost of goods",
     f"=-({names['c1']}+{names['c2']}+{names['c3']}+{names['c4']})"),
    ("Less secondary freight", f"=-{names['frt']}"),
    ("Less trade spend", f"=-{names['trd']}"),
    ("CONTRIBUTION PER SACHET", "=B8+B9+B10+B11"),
    ("Margin on what we receive", "=B12/B8"),
]
r = 5
for label, formula in unit:
    u.cell(r, 1, label)
    c = u.cell(r, 2, formula)
    c.number_format = "0.00%" if "Margin" in label else "0.00"
    c.border = BOX
    if "WE RECEIVE" in label or "CONTRIBUTION" in label:
        u.cell(r, 1).font = BOLD
        c.font = BOLD
    r += 1

u["A15"] = "ANNUAL, AT THE PLANNING CASE"
u["A15"].font = Font(bold=True, color=ORANGE)
ann = [
    ("Sachets per month", f"={names['out']}*{names['rate']}*{names['days']}", "#,##0"),
    ("Contribution per month, Rs", "=B16*B12", "#,##0"),
    ("Contribution per year, Rs", "=B17*12", "#,##0"),
    ("Contribution per year, GBP", f"=B18/{names['fx']}", "£#,##0"),
]
r = 16
for label, formula, fmt in ann:
    u.cell(r, 1, label)
    c = u.cell(r, 2, formula)
    c.number_format = fmt
    c.border = BOX
    if "GBP" in label:
        u.cell(r, 1).font = BOLD
        c.font = BOLD
    r += 1
u.column_dimensions["A"].width = 40
u.column_dimensions["B"].width = 16

# ---------------------------------------------------------------- SCENARIOS
s = wb.create_sheet("Scenarios")
title(s, "DOWNSIDE — MARGIN CASES CROSSED WITH VOLUME",
      "Flexes only what Raemy named: cost of goods, contribution margin, distribution costs.")

s["A4"] = "Case"
for i, h in enumerate(["COGS Rs", "Retailer %", "Distributor %", "Freight", "Trade",
                       "We get", "Contribution", "Margin"], start=2):
    s.cell(4, i, h)
for i in range(1, 10):
    s.cell(4, i).font = BOLD
    s.cell(4, i).fill = HEAD_FILL
    s.cell(4, i).border = BOX

cases = [
    ("1. Plan, as pitched", 2.10, 0.15, 0.08, 0.15, 0.30),
    ("2. Cost of goods only", 2.75, 0.15, 0.08, 0.15, 0.30),
    ("3. Trade terms only", 2.10, 0.18, 0.10, 0.20, 0.50),
    ("4. Both — the case I would plan on", 2.75, 0.18, 0.10, 0.20, 0.50),
    ("5. Stress", 3.25, 0.20, 0.12, 0.25, 0.60),
]
r = 5
for label, cogs, ret, dis, frt, trd in cases:
    s.cell(r, 1, label)
    for i, v in enumerate([cogs, ret, dis, frt, trd], start=2):
        c = s.cell(r, i, v)
        c.fill = INPUT_FILL
        c.border = BOX
        c.number_format = "0.00%" if i in (3, 4) else "0.00"
    s.cell(r, 7, f"={names['mrp']}/(1+{names['gst']})*(1-C{r})*(1-D{r})").number_format = "0.00"
    s.cell(r, 8, f"=G{r}-B{r}-E{r}-F{r}").number_format = "0.00"
    s.cell(r, 9, f"=H{r}/G{r}").number_format = "0.0%"
    for i in (7, 8, 9):
        s.cell(r, i).border = BOX
    if label.startswith("4."):
        for i in range(1, 10):
            s.cell(r, i).font = BOLD
    r += 1

s["A12"] = "ANNUAL CONTRIBUTION, GBP m — volume down the side, margin case across"
s["A12"].font = Font(bold=True, color=ORANGE)
s["A13"] = "Sachets/month"
for i in range(1, 6):
    s.cell(13, i + 1, f"Case {i}")
    s.cell(13, i + 1).font = BOLD
    s.cell(13, i + 1).fill = HEAD_FILL
    s.cell(13, i + 1).border = BOX
s["A13"].font = BOLD
s["A13"].fill = HEAD_FILL

vols = [("Bull, 10/outlet/day", f"={names['out']}*10*{names['days']}"),
        ("Base, 2.5/day (plan)", f"={names['out']}*{names['rate']}*{names['days']}"),
        ("Low, 1.5/day", f"={names['out']}*1.5*{names['days']}")]
r = 14
for label, vol in vols:
    s.cell(r, 1, label).font = BOLD if "Base" in label else Font()
    for i in range(1, 6):
        s.cell(r, i + 1,
               f"=({vol})*H{4+i}*12/{names['fx']}/1000000").number_format = "0.00"
        s.cell(r, i + 1).border = BOX
    r += 1

s["A18"] = ("The honest planning number is the Base row, case 4. Every figure above is a formula: "
            "change a yellow cell and the whole grid moves.")
s["A18"].font = Font(italic=True, color="6F685C")
s.column_dimensions["A"].width = 36
for col in "BCDEFGHI":
    s.column_dimensions[col].width = 13

# ---------------------------------------------------------------- SETUP
p = wb.create_sheet("Setup")
title(p, "ONE-OFF COST TO GET GOING, GBP",
      "7 of 14 lines are still requests for quotation rather than quotes.")
p["A4"], p["B4"], p["C4"] = "Line", "GBP", "Status"
for c in ("A4", "B4", "C4"):
    p[c].font = BOLD
    p[c].fill = HEAD_FILL
    p[c].border = BOX

setup = [
    ("Indian company: incorporation as a WOS, FDI/FEMA filings", 3500, "TARGET — KNM quoted 6,627, rejected as overscoped"),
    ("Resident director service, 12 months", 2666, "KNM 11 Aug, USD 3,600/yr. Fair, pay it"),
    ("FSSAI central licence", 1000, "RFQ — nobody has quoted this"),
    ("Accounting, statutory audit, company secretarial, 12 months", 5000, "TARGET — KNM quoted 13,481, rejected"),
    ("Trade mark: clearance search and filing, 3 classes", 759, "Intepat, instructed 9 Aug"),
    ("Trade mark: prosecution to registration certificate", 600, "RFQ"),
    ("Formula and flavour development, blind sip tests", 6000, ""),
    ("Lab work: nutritional analysis, stability at 45C/75% RH", 2000, "RFQ"),
    ("Production-ready artwork: sachet, outer box, code layout", 5000, ""),
    ("Print cylinders and laminate setup", 2000, "RFQ"),
    ("First production run, 500,000 sachets at Rs 3.25", 12626, "RFQ — sized by the factory's minimum, not the pilot"),
    ("Distributor trade scheme and incentives, 90 days", 2000, ""),
    ("India ground support, part time, 3 months", 5000, ""),
    ("Travel, two trips", 5000, ""),
    ("Unique-code system and retailer app", 0, "Built in house by Felix, Jesse 11 Aug"),
]
r = 5
for label, val, note in setup:
    p.cell(r, 1, label)
    c = p.cell(r, 2, val)
    c.number_format = "£#,##0"
    c.fill = INPUT_FILL
    c.border = BOX
    p.cell(r, 3, note).font = Font(color="6F685C", size=9)
    r += 1
p.cell(r, 1, "Subtotal").font = BOLD
p.cell(r, 2, f"=SUM(B5:B{r-1})").number_format = "£#,##0"
p.cell(r, 2).font = BOLD
p.cell(r + 1, 1, "Contingency at 15%")
p.cell(r + 1, 2, f"=B{r}*0.15").number_format = "£#,##0"
p.cell(r + 2, 1, "TOTAL").font = WHITE_BOLD
p.cell(r + 2, 1).fill = PatternFill("solid", fgColor=ORANGE)
p.cell(r + 2, 2, f"=B{r}+B{r+1}").number_format = "£#,##0"
p.cell(r + 2, 2).font = WHITE_BOLD
p.cell(r + 2, 2).fill = PatternFill("solid", fgColor=ORANGE)
p.cell(r + 4, 1, "Working capital note: the first run is sized by the factory's minimum order, not "
                 "by the pilot. At 2.5 sachets/outlet/day a 300-outlet 90-day pilot consumes about "
                 "67,500 sachets, so roughly 432,500 sachets sit in stock when it ends.")
p.cell(r + 4, 1).font = Font(italic=True, color="6F685C")
p.column_dimensions["A"].width = 56
p.column_dimensions["B"].width = 14
p.column_dimensions["C"].width = 56

out = "/Users/bgvai/agents/junior/workspace/docs/everest-power/BGEP-Financial-Model.xlsx"
wb.save(out)
print("saved:", out)
